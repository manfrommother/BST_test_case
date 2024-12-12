from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db.models import Count
from django.utils import timezone
from openpyxl import Workbook
import json
from datetime import datetime, timedelta
from .models import Robot

@method_decorator(csrf_exempt, name='dispatch')
class RobotCreateView(View):
    """Представление для создания робота через API"""
    
    def post(self, request, *args, **kwargs):
        """Обработка POST-запроса для создания робота"""
        try:
            data = json.loads(request.body)
            
            # Проверяем длину модели
            if len(data.get('model', '')) != 2:
                return JsonResponse(
                    {'error': 'Model must be exactly 2 characters long'}, 
                    status=400
                )
                
            # Проверяем длину версии
            if len(data.get('version', '')) != 2:
                return JsonResponse(
                    {'error': 'Version must be exactly 2 characters long'}, 
                    status=400
                )
                
            # Проверяем дату
            try:
                created = datetime.strptime(data['created'], "%Y-%m-%d %H:%M:%S")
            except (ValueError, KeyError):
                return JsonResponse(
                    {'error': 'Invalid datetime format. Use YYYY-MM-DD HH:MM:SS'}, 
                    status=400
                )
            
            # Создаем робота
            robot = Robot.objects.create(
                model=data['model'],
                version=data['version'],
                created=created
            )
            
            return JsonResponse({
                'model': robot.model,
                'version': robot.version,
                'created': robot.created.strftime("%Y-%m-%d %H:%M:%S")
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse(
                {'error': 'Invalid JSON format'}, 
                status=400
            )
        except Exception as e:
            return JsonResponse(
                {'error': str(e)}, 
                status=500
            )
        

class RobotExcelReportView(View):
    """Представление для генерации Excel-отчета по роботам"""
    
    def get_last_week_data(self):
        """Получаем данные за последнюю неделю"""
        end_date = timezone.now()
        start_date = end_date - timedelta(days=7)
        
        return Robot.objects.filter(
            created__range=(start_date, end_date)
        ).values(
            'model', 'version'
        ).annotate(
            count=Count('id')
        ).order_by('model', 'version')

    def get(self, request, *args, **kwargs):
        """Обработка GET-запроса для скачивания отчета"""
        # Создаем новый Excel-файл
        wb = Workbook()
        
        # Получаем данные за последнюю неделю
        robots_data = self.get_last_week_data()
        
        # Группируем данные по моделям
        models_data = {}
        for item in robots_data:
            if item['model'] not in models_data:
                models_data[item['model']] = []
            models_data[item['model']].append(item)
        
        # Удаляем стандартный лист
        if len(models_data) > 0:
            wb.remove(wb.active)
        
        # Создаем лист для каждой модели
        for model, data in models_data.items():
            # Создаем новый лист с названием модели
            ws = wb.create_sheet(title=f"Model {model}")
            
            # Добавляем заголовки
            ws['A1'] = 'Модель'
            ws['B1'] = 'Версия'
            ws['C1'] = 'Количество за неделю'
            
            # Заполняем данные
            row = 2
            for item in data:
                ws[f'A{row}'] = item['model']
                ws[f'B{row}'] = item['version']
                ws[f'C{row}'] = item['count']
                row += 1
            
            # Устанавливаем ширину столбцов
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
        else:
            # Если данных нет, оставляем один лист с сообщением
            ws = wb.active
            ws['A1'] = 'Нет данных за последнюю неделю'
            ws.column_dimensions['A'].width = 30

        # Формируем имя файла с текущей датой
        filename = f'robots_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
        
        # Создаем HTTP-ответ с Excel-файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Сохраняем файл в response
        wb.save(response)
        
        return response

